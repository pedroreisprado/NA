import pandas as pd
from PyPDF2 import PdfReader,PdfWriter
import os 
import re
from datetime import datetime
from openpyxl import load_workbook
from config import SAVE_PATH,NEW_PATH,DADOS_PATH,DEVOLUTIVAS

def separar_paginas(pdf_path):
    pdf_path = pdf_path.replace('.pdf','')
    pdf_path = (f'{SAVE_PATH}\\{pdf_path}.pdf')
    # Crie um diretório para armazenar os PDFs separados

    output_dir = (f'{NEW_PATH}\\pdf_apartado')
    os.makedirs(output_dir, exist_ok=True)

    with open(pdf_path, "rb") as file:
        pdf_reader = PdfReader(file)
        total_pages = len(pdf_reader.pages)

        # Para cada página, crie um PDF separado
        for page_number in range(total_pages):
            pdf_writer = PdfWriter()
            pdf_writer.add_page(pdf_reader._get_page(page_number))
            page_text = pdf_reader._get_page(page_number).extract_text()

            # Buscar pelo texto "unidade"
            index_unidade = page_text.find("Unidade: ")
            if index_unidade != -1:
                # Pegar os dois caracteres após "unidade"
                codigo_unidade = page_text[index_unidade + len("Unidade: "): index_unidade + len("Unidade: ") + 2]
                print(f"Código da unidade na página {page_number + 1}: {codigo_unidade}")

            # Salvar a página atual como um PDF separado
            output_path = os.path.join(output_dir, f"{codigo_unidade}.pdf")
            with open(output_path, "wb") as output_file:
                pdf_writer.write(output_file)


    files = os.listdir(f'{NEW_PATH}\\pdf_apartado')
    return files

def getData(unidade,operacao):
    unidade = str(unidade)
    df = pd.read_excel(f'{DADOS_PATH}//unidpl01.xls', skiprows= 8)
    # Encontre o índice da linha onde o valor '10' está na coluna 0
    indice = df[df.iloc[:, 0] == unidade].index[0]

    # Encontre o valor na coluna 8 na mesma linha do valor '10'
    if operacao == "Inquilino":
        valor_na_coluna_8 = df.iloc[indice + 3, 6]
        nome = df.iloc[indice + 3, 1]

        if 'Unidade Vazia' in nome:
            valor_na_coluna_8 = df.iloc[indice + 2, 6]
            nome = df.iloc[indice + 2, 1]

        if pd.isna(valor_na_coluna_8):
            print('Não foi possivel localizar o WPP do IQ, passando para o WPP do PP')
            valor_na_coluna_8 = df.iloc[indice + 2, 6]
            nome = df.iloc[indice + 2, 1]
            if pd.isna(valor_na_coluna_8):
                return False
        return [valor_na_coluna_8,nome]

    valor_na_coluna_8 = df.iloc[indice + 2, 6]
    nome = df.iloc[indice + 2, 1]

    if pd.isna(valor_na_coluna_8):
        return False
    return [valor_na_coluna_8,nome]

def deletePDF():
    path = (f'{NEW_PATH}\\pdf_apartado')
    pdfs = os.listdir(path)
    print(pdfs)
    for file in pdfs:
        caminho_arquivo = os.path.join(path, file)
        os.remove(caminho_arquivo)
    return True

def CreateSheet(condominio,operacao):
    data_hoje = str(datetime.today().strftime("%d-%m-%Y"))
    #-- Abrindo o XLSX de devolutivas
    workbook = load_workbook(DEVOLUTIVAS)
    #-- Criando aba do dia
    sheet_start = workbook['start']
    last_row = sheet_start.max_row+1
    sheet_start.cell(row=last_row, column=1).value = data_hoje
    sheet_start.cell(row=last_row, column=3).value = condominio
    sheet_start.cell(row=last_row, column=2).value = operacao

    workbook.create_sheet(data_hoje)
    
    sheet = workbook[data_hoje]
    sheet.cell(row=1, column=1).value = 'Data'
    sheet.cell(row=1, column=2).value = 'Condominio'
    sheet.cell(row=1, column=3).value = 'Operação'
    sheet.cell(row=1, column=4).value = 'Nome'
    sheet.cell(row=1, column=5).value = 'Unidade'
    sheet.cell(row=1, column=6).value = 'Telefone'
    workbook.save(DEVOLUTIVAS)
    return True

def insertInfo(nome, unidade, condominio, telefone, operacao):
    #-- Obter a data de hoje
    data_hoje = str(datetime.today().strftime("%d-%m-%Y"))
    #-- Abrindo o XLSX de devolutivas
    workbook = load_workbook(DEVOLUTIVAS)

    sheet = workbook[data_hoje]
    #-- Coletando o numero da ultima linha preenchida
    last_row = sheet.max_row +1

    #-- Inputando informação do cliente nas devolutivas
    sheet.cell(row=last_row, column=1).value = data_hoje
    sheet.cell(row=last_row, column=2).value = condominio
    sheet.cell(row=last_row, column=3).value = operacao
    sheet.cell(row=last_row, column=4).value = nome
    sheet.cell(row=last_row, column=5).value = unidade
    sheet.cell(row=last_row, column=6).value = telefone
    workbook.save(DEVOLUTIVAS)
    return True
